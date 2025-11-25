import streamlit as st
import requests
from io import BytesIO
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import re

st.set_page_config(page_title="Insertar imágenes en Excel", layout="centered")

st.title("📸 Insertar imágenes en Excel desde URLs (múltiples por celda)")
st.write(
    "Sube un archivo Excel que tenga una columna llamada "
    "**'Imágenes del anaquel AL LLEGAR'** con uno o varios links de imágenes por celda.\n\n"
    "Si hay varias URLs en una misma celda, se insertarán varias imágenes en la misma fila, "
    "usando columnas hacia la derecha (F, G, H, …)."
)

uploaded_file = st.file_uploader("Subir archivo Excel", type=["xlsx"])

if uploaded_file is not None:
    if st.button("🔄 Procesar archivo e insertar imágenes"):
        with st.spinner("Procesando archivo, descargando imágenes..."):
            try:
                # Cargar workbook desde el archivo subido
                file_bytes = BytesIO(uploaded_file.read())
                wb = openpyxl.load_workbook(file_bytes)
                ws = wb.active  # o wb["NombreHoja"] si quieres una hoja específica

                # Encontrar la columna que tenga el encabezado deseado
                header_name = "Imágenes del anaquel AL LLEGAR"
                image_col_idx = None

                for cell in ws[1]:  # primera fila = encabezados
                    if str(cell.value).strip() == header_name:
                        image_col_idx = cell.column
                        break

                if image_col_idx is None:
                    st.error(
                        f"No se encontró la columna con el encabezado: '{header_name}'.\n"
                        "Verifica que el nombre coincida exactamente."
                    )
                else:
                    max_row = ws.max_row
                    progress = st.progress(0)
                    processed = 0
                    total = max_row - 1 if max_row > 1 else 1

                    # Recorremos filas de datos
                    for row in range(2, max_row + 1):
                        cell = ws.cell(row=row, column=image_col_idx)
                        url_text = str(cell.value).strip() if cell.value is not None else ""

                        if url_text:
                            # Extraer TODAS las URLs de la celda
                            urls = re.findall(r"https?://\S+", url_text)

                            for idx, img_url in enumerate(urls):
                                try:
                                    resp = requests.get(img_url, timeout=10)
                                    if resp.status_code == 200:
                                        img_bytes = BytesIO(resp.content)
                                        img = XLImage(img_bytes)

                                        # Tamaño de miniatura (ajusta a tu gusto)
                                        img.width = 100
                                        img.height = 100

                                        # Columna para esta imagen:
                                        # primera imagen en la columna original (F),
                                        # segunda en la siguiente (G), tercera en H, etc.
                                        col_idx = image_col_idx + idx
                                        col_letter = get_column_letter(col_idx)
                                        anchor_cell = f"{col_letter}{row}"

                                        img.anchor = anchor_cell
                                        ws.add_image(img)

                                        # Subir un poco el ancho de la columna
                                        if ws.column_dimensions[col_letter].width is None or \
                                           ws.column_dimensions[col_letter].width < 18:
                                            ws.column_dimensions[col_letter].width = 18

                                        # Ajustar altura de la fila para que quepan bien las imágenes
                                        if ws.row_dimensions[row].height is None or \
                                           ws.row_dimensions[row].height < 80:
                                            ws.row_dimensions[row].height = 80
                                    else:
                                        # Si no descarga bien, solo la saltamos
                                        pass
                                except Exception:
                                    # Si falla alguna imagen, no detenemos el proceso
                                    pass

                        processed += 1
                        progress.progress(processed / total)

                    # Guardar workbook en memoria
                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)

                    st.success("¡Proceso terminado! Ya puedes descargar tu archivo con todas las imágenes.")
                    st.download_button(
                        label="⬇️ Descargar Excel con imágenes",
                        data=output,
                        file_name="excel_con_imagenes.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

            except Exception as e:
                st.error(f"Ocurrió un error procesando el archivo: {e}")
else:
    st.info("Sube un archivo Excel para comenzar.")
